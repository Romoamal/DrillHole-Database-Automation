import matplotlib
matplotlib.use('Agg')  # Use non-GUI backend for Matplotlib

import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from flask import Flask, request, send_file, render_template_string
import openpyxl
import threading
import os
import pandas as pd
import logging
from datetime import datetime
import matplotlib.pyplot as plt
from io import BytesIO
import base64

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Configuration
daily_folder = r'Daily_Data'
database_file = r'C:\Users\CUDY\DrillHole_Automation\drilling_database.xlsx'

class Watcher:
    def __init__(self):
        self.observer = Observer()

    def run(self):
        event_handler = Handler()
        self.observer.schedule(event_handler, daily_folder, recursive=False)
        self.observer.start()
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            self.observer.stop()
        self.observer.join()

class Handler(FileSystemEventHandler):
    def on_created(self, event):
        if event.is_directory:
            return
        if event.src_path.endswith('.xls') or event.src_path.endswith('.xlsx'):
            logger.info(f"New file detected: {event.src_path}")
            try:
                process_new_file(event.src_path)
                logger.info(f"Successfully processed file: {event.src_path}")
            except Exception as e:
                logger.error(f"Error processing file {event.src_path}: {str(e)}")

def process_new_file(file_path):
    time.sleep(2)  # Ensure file is fully written
    transformed_data = transform_data(file_path)
    append_to_database(transformed_data)

def transform_data(file_path):
    try:
        # Load workbook for metadata
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        raw_hole_id = sheet['B3'].value
        raw_date_logging = sheet['L4'].value
        logger.info(f"Raw metadata from {os.path.basename(file_path)} - B3: {raw_hole_id}, L4: {raw_date_logging}")

        hole_id = str(raw_hole_id).strip() if raw_hole_id else None
        date_logging = raw_date_logging
        if isinstance(date_logging, datetime):
            date_logging = date_logging.date()
        elif isinstance(date_logging, str):
            try:
                date_logging = datetime.strptime(date_logging, '%Y-%m-%d %H:%M:%S').date()
            except ValueError:
                date_logging = None
        
        if not hole_id or not date_logging:
            raise ValueError(f"Missing or invalid Hole ID ({hole_id}) or Date Logging ({date_logging}) in metadata")
        # Read data section with headers from row 8
        daily_data = pd.read_excel(
            file_path,
            skiprows=5,  # Skip to row 5 (0-based index)
            header=0,   # Use row 6 as headers
            usecols='B:V'
        )

        # Log column names for debugging
        logger.info(f"Columns in {os.path.basename(file_path)}: {list(daily_data.columns)}")

        # Check if expected columns exist
        required_columns = ['FROM', 'TO', 'INTERVAL (M)', 'ACT CORE (M)', 'RECOVERY (%)',
                            'GENERAL LITHOLOGY', 'SUB GEN LITHOLOGY', 'ROCK CODE', 'GRAIN SIZE',
                            'WEATHERING', 'COLOUR PRIMARY', 'MINERALS PRIMARY', 'MINERALS SECONDARY', 'MINERALS TERTIARY']
        missing_columns = [col for col in required_columns if col not in daily_data.columns]
        if missing_columns:
            logger.warning(f"Missing expected columns: {missing_columns}. Falling back to indices.")
            # Fallback to indices (assuming B:V order matches C06-090.xlsx)
            daily_data = pd.read_excel(
                file_path,
                skiprows=6,  # Data starts at row 7
                header=None,
                usecols='B:V'
            )
            transformed_rows = []
            for _, row in daily_data.iterrows():
                if pd.isna(row[0]):  # FROM (B)
                    continue
                transformed_row = {
                    'Date Logging': date_logging,
                    'Hole ID': hole_id,
                    'From': row[0],  # B
                    'To': row[1],  # C
                    'Length': row[2],  # D
                    'Actual Core': row[3],  # E
                    'Recovery pecentage': row[6] / 100 if pd.notna(row[6]) else 1.0,  # H
                    'Material Code': str(row[7]).lower() if pd.notna(row[7]) else '',  # I
                    'Layer Code': str(row[8]).lower() if pd.notna(row[8]) else '',  # J
                    'Rock Code': str(row[9]).lower() if pd.notna(row[9]) else '',  # K
                    'Grain': str(row[10]).lower() if pd.notna(row[10]) else '',  # L
                    'Weath': row[12] if pd.notna(row[12]) else None,  # N
                    'Colour': str(row[13]).lower() if pd.notna(row[13]) else '',  # O
                    'Minerals Pri': str(row[16]).lower() if pd.notna(row[16]) else '',  # R
                    'Minerals Sec': str(row[17]).lower() if pd.notna(row[17]) else '',  # S
                    'Minerals Ter': str(row[18]).lower() if pd.notna(row[18]) else '',  # T
                    'Bolder leght (m)': None
                }
                transformed_rows.append(transformed_row)
        else:
            # Use column names if available
            transformed_rows = []
            for _, row in daily_data.iterrows():
                if pd.isna(row['FROM']):
                    continue
                transformed_row = {
                    'Date Logging': date_logging,
                    'Hole ID': hole_id,
                    'From': row['FROM'],
                    'To': row['TO'],
                    'Length': row['INTERVAL (M)'],
                    'Actual Core': row['ACT CORE (M)'],
                    'Recovery pecentage': row['RECOVERY (%)'] / 100 if pd.notna(row['RECOVERY (%)']) else 1.0,
                    'Material Code': str(row['GENERAL LITHOLOGY']).lower() if pd.notna(row['GENERAL LITHOLOGY']) else '',
                    'Layer Code': str(row['SUB GEN LITHOLOGY']).lower() if pd.notna(row['SUB GEN LITHOLOGY']) else '',
                    'Rock Code': str(row['ROCK CODE']).lower() if pd.notna(row['ROCK CODE']) else '',
                    'Grain': str(row['GRAIN SIZE']).lower() if pd.notna(row['GRAIN SIZE']) else '',
                    'Weath': row['WEATHERING'] if pd.notna(row['WEATHERING']) else None,
                    'Colour': str(row['COLOUR PRIMARY']).lower() if pd.notna(row['COLOUR PRIMARY']) else '',
                    'Minerals Pri': str(row['MINERALS PRIMARY']).lower() if pd.notna(row['MINERALS PRIMARY']) else '',
                    'Minerals Sec': str(row['MINERALS SECONDARY']).lower() if pd.notna(row['MINERALS SECONDARY']) else '',
                    'Minerals Ter': str(row['MINERALS TERTIARY']).lower() if pd.notna(row['MINERALS TERTIARY']) else '',
                    'Bolder leght (m)': None
                }
                transformed_rows.append(transformed_row)
        
        return transformed_rows
        
    except Exception as e:
        logger.error(f"Error in transform_data: {str(e)}")
        raise

def append_to_database(transformed_rows):
    try:
        db_path = os.path.normpath(database_file)
        logger.info(f"Attempting to update database at: {db_path}")
        
        wb = openpyxl.load_workbook(db_path)
        sheet = wb.active
        last_row = sheet.max_row
        
        if last_row == 1 and sheet.cell(row=1, column=1).value is None:
            headers = [
                'Date Logging', 'Hole ID', 'From', 'To', 'Length', 'Actual Core', 'Recovery pecentage',
                'Material Code', 'Layer Code', 'Rock Code', 'Grain', 'Weath', 'Colour',
                'Minerals Pri', 'Minerals Sec', 'Minerals Ter', 'Bolder leght (m)'
            ]
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col).value = header
            last_row = 1

        for row_data in transformed_rows:
            last_row += 1
            sheet.cell(row=last_row, column=2).value = row_data['Date Logging']
            sheet.cell(row=last_row, column=3).value = row_data['Hole ID']
            sheet.cell(row=last_row, column=4).value = row_data['From']
            sheet.cell(row=last_row, column=5).value = row_data['To']
            sheet.cell(row=last_row, column=6).value = row_data['Length']
            sheet.cell(row=last_row, column=7).value = row_data['Actual Core']
            sheet.cell(row=last_row, column=8).value = row_data['Recovery pecentage']
            sheet.cell(row=last_row, column=9).value = row_data['Material Code']
            sheet.cell(row=last_row, column=10).value = row_data['Layer Code']
            sheet.cell(row=last_row, column=11).value = row_data['Rock Code']
            sheet.cell(row=last_row, column=12).value = row_data['Grain']
            sheet.cell(row=last_row, column=14).value = row_data['Weath']
            sheet.cell(row=last_row, column=15).value = row_data['Colour']
            sheet.cell(row=last_row, column=18).value = row_data['Minerals Pri']
            sheet.cell(row=last_row, column=19).value = row_data['Minerals Sec']
            sheet.cell(row=last_row, column=20).value = row_data['Minerals Ter']
            sheet.cell(row=last_row, column=22).value = row_data['Bolder leght (m)']
        
        wb.save(db_path)
        logger.info(f"Successfully appended {len(transformed_rows)} rows to {db_path}")
        
    except Exception as e:
        logger.error(f"Error appending to database: {str(e)}")
        raise

# Helper function to load database
def load_database():
    return pd.read_excel(database_file)

# Visualization functions
def generate_recovery_plot(df):
    # Group by depth (average 'From' or 'To') and calculate average recovery
    depth_avg = df.groupby('From')['Recovery pecentage'].mean().reset_index()
    if depth_avg.empty:
        return None
    plt.figure(figsize=(10, 6))
    plt.plot(depth_avg['From'], depth_avg['Recovery pecentage'], marker='o', linestyle='-', color='blue')
    plt.title('Average Recovery Percentage vs. Average Depth')
    plt.xlabel('Average Depth (m) - From')
    plt.ylabel('Average Recovery Percentage')
    plt.grid(True)
    img = BytesIO()
    plt.savefig(img, format='png', bbox_inches='tight')
    plt.close()
    img.seek(0)
    return base64.b64encode(img.getvalue()).decode('utf-8')

def generate_material_distribution_plot(df):
    material_counts = df['Material Code'].value_counts()
    plt.figure(figsize=(10, 6))
    material_counts.plot(kind='bar', color='purple')
    plt.title('Distribution of Material Codes')
    plt.xlabel('Material Code')
    plt.ylabel('Count')
    plt.xticks(rotation=45)
    plt.grid(axis='y')
    img = BytesIO()
    plt.savefig(img, format='png', bbox_inches='tight')
    plt.close()
    img.seek(0)
    return base64.b64encode(img.getvalue()).decode('utf-8')

# Dashboard HTML template
DASHBOARD_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>DrillHole Analytics Dashboard</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 0; background-color: #f5f7fa; color: #333; }
        .container { max-width: 1200px; margin: 0 auto; padding: 20px; }
        .sidebar { width: 200px; float: left; background-color: #e9ecef; padding: 20px; height: 100vh; position: fixed; }
        .content { margin-left: 220px; }
        h1 { text-align: center; color: #2c3e50; margin-bottom: 20px; }
        .stats-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 20px; margin-bottom: 20px; }
        .stat-card { background-color: white; border-radius: 8px; padding: 15px; text-align: center; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
        .stat-card h3 { margin: 0 0 5px 0; font-size: 14px; color: #7f8c8d; }
        .stat-card p { margin: 0; font-size: 18px; color: #2c3e50; }
        .plot-section { background-color: white; border-radius: 8px; padding: 20px; margin-bottom: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
        .plot { text-align: center; }
        img { max-width: 100%; height: auto; }
        .form-container { text-align: center; margin-bottom: 20px; }
        select, button { padding: 8px; font-size: 14px; border-radius: 4px; border: 1px solid #ccc; }
        button { background-color: #3498db; color: white; cursor: pointer; }
        button:hover { background-color: #2980b9; }
    </style>
</head>
<body>
    <div class="sidebar">
        <h2>Navigation</h2>
        <ul style="list-style-type: none; padding: 0;">
            <li><a href="/" style="color: #2c3e50; text-decoration: none;">Dashboard</a></li>
            <li><a href="/update" style="color: #2c3e50; text-decoration: none;">Update Data</a></li>
            <li><a href="/database-status" style="color: #2c3e50; text-decoration: none;">Database Status</a></li>
        </ul>
    </div>
    <div class="content">
        <div class="container">
            <h1>DrillHole Analytics Dashboard</h1>
            <div class="stats-grid">
                <div class="stat-card">
                    <h3>Total Drill Holes</h3>
                    <p>{{ total_holes }}</p>
                </div>
                <div class="stat-card">
                    <h3>Average Depth</h3>
                    <p>{{ avg_depth|round(2) }} m</p>
                </div>
                <div class="stat-card">
                    <h3>Deepest Depth</h3>
                    <p>{{ deepest_depth|round(2) }} m</p>
                </div>
                <div class="stat-card">
                    <h3>Shallowest Depth</h3>
                    <p>{{ shallowest_depth|round(2) }} m</p>
                </div>
            </div>
            <div class="plot-section">
                <h2>Average Recovery Percentage vs. Average Depth</h2>
                <div class="plot">
                    {% if recovery_plot %}
                        <img src="data:image/png;base64,{{ recovery_plot }}" alt="Recovery Plot">
                    {% else %}
                        <p>No data available.</p>
                    {% endif %}
                </div>
            </div>
            <div class="plot-section">
                <h2>Material Code Distribution</h2>
                <div class="plot">
                    <img src="data:image/png;base64,{{ material_plot }}" alt="Material Distribution">
                </div>
            </div>
        </div>
    </div>
</body>
</html>
"""

# Flask Routes
@app.route('/', methods=['GET'])
def dashboard():
    try:
        df = load_database()
        hole_ids = df['Hole ID'].unique().tolist()

        # General statistics
        total_holes = len(hole_ids) if hole_ids else 0
        avg_depth = df['Length'].mean() if not df.empty else 0
        deepest_depth = df['To'].max() if not df.empty else 0
        shallowest_depth = df['From'].min() if not df.empty else 0

        recovery_plot = generate_recovery_plot(df)
        material_plot = generate_material_distribution_plot(df)

        return render_template_string(
            DASHBOARD_TEMPLATE,
            total_holes=total_holes,
            avg_depth=avg_depth,
            deepest_depth=deepest_depth,
            shallowest_depth=shallowest_depth,
            hole_ids=hole_ids,
            recovery_plot=recovery_plot,
            material_plot=material_plot
        )
    except Exception as e:
        logger.error(f"Error rendering dashboard: {str(e)}")
        return f"Error: {str(e)}", 500

@app.route('/update', methods=['GET', 'POST'])
def trigger_update():
    try:
        if request.method == 'POST':
            processed_files = []
            for file in os.scandir(daily_folder):
                if file.is_file() and (file.name.endswith('.xls') or file.name.endswith('.xlsx')):
                    process_new_file(file.path)
                    processed_files.append(file.name)
            return {'status': 'success', 'processed_files': processed_files}
        else:  # GET request
            processed_files = []
            for file in os.scandir(daily_folder):
                if file.is_file() and (file.name.endswith('.xls') or file.name.endswith('.xlsx')):
                    process_new_file(file.path)
                    processed_files.append(file.name)
            return {'status': 'success', 'processed_files': processed_files, 'message': 'Manual update triggered via GET'}
    except Exception as e:
        return {'status': 'error', 'message': str(e)}, 500

@app.route('/database-status', methods=['GET'])
def database_status():
    try:
        wb = openpyxl.load_workbook(database_file)
        sheet = wb.active
        return {'status': 'success', 'last_row': sheet.max_row}, 200
    except Exception as e:
        return {'status': 'error', 'message': str(e)}, 500

if __name__ == '__main__':
    watcher_thread = threading.Thread(target=lambda: Watcher().run())
    watcher_thread.daemon = True
    watcher_thread.start()
    app.run(host='127.0.0.1', port=5000)